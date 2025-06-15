# 엑셀 파일을 생성하기 위해서 openpyxl의 Workbook 클래스
from openpyxl import Workbook

# 테두리 스타일을 위한 클래스
from openpyxl.styles import Border, Side

# 폰트 스타일을 위한 클래스
from openpyxl.styles import Font

# 텍스트 정렬을 위한 클래스
from openpyxl.styles import Alignment

# 학생 수 입력
n = int(input("학생 수를 입력해주세요 : "))

# 폰트 종류
# name-폰트 이름 / size-폰트 크기 / bold-굵기 / italic-기울기 / underline-밑줄
Title_font = Font(name='Pretendard', size=24, bold=True)
Pretendard = Font(name='Pretendard', size=12, bold=True)

# 얇은 테두리
Thin_border = Border(
    left = Side(style = 'thin'),
    right = Side(style = 'thin'),
    top = Side(style = 'thin'),
    bottom = Side(style = 'thin')
)

# 굵은 테두리
Thick_border = Side(style = 'thick')

# 테두리 없음
No_border = Side(style = None)

# 새 Workbook(엑셀 파일) 생성
xlsx = Workbook()

# 첫번째 워크시트 활성화
x1 = xlsx.active

# 전체 범위 지정
min_row, max_row = 1, 26 # 행
min_col, max_col = 1, 13  # 열

# 각 셀에 대해 위치에 따라 테두리 적용
for row in range(min_row, max_row + 1):
    for col in range(min_col, max_col + 1):
        cell = x1.cell(row = row, column = col)

        # 각 방향 테두리 조건부 설정
        if row == min_row:
            top = Thick_border
            bottom = No_border
        elif row == max_row:
            top = No_border
            bottom = Thick_border
        else:
            top = No_border
            bottom = No_border

        if col == min_col:
            left = Thick_border
            right = No_border
        elif col == max_col:
            left = No_border
            right = Thick_border
        else:
            left = No_border
            right = No_border

        cell.border = Border(top = top, bottom = bottom, left = left, right = right)
        cell.alignment = Alignment(horizontal='center', vertical='center')

# 기본 병합
x1.merge_cells('B2:L3') # 좌석 배정표
x1.merge_cells('B22:C22') # 학반
x1.merge_cells('B23:C24') # 담임선생님
x1.merge_cells('E23:I24') # 칠판

# 기본 테두리
for row in x1['B2:L3']: # 좌석 배치표(타이틀)
    for cell in row:
        cell.border = Thin_border

for row in x1['B22:C22']: # 학반
    for cell in row:
        cell.border = Thin_border

for row in x1['B23:C24']: # 담임 선생님
    for cell in row:
        cell.border = Thin_border

for row in x1['E23:I24']: # 칠판
    for cell in row:
        cell.border = Thin_border

# 기본 데이터 입력
x1['B2'] = "좌석 배정표"
x1['B2'].font = Title_font

x1['E23'] = "칠판"
x1['E23'].font = Pretendard

# 자리표 배치
current_seat = 1  # 최소 1자리 이상

# 자리 배치를 위한 좌표 리스트
seat_positions = [
    ('K17:L19', 17, 11),  # 1
    ('H17:I19', 17, 8),   # 2
    ('E17:F19', 17, 5),   # 3
    ('B17:C19', 17, 2),   # 4
    ('K14:L16', 14, 11),  # 5
    ('H14:I16', 14, 8),   # 6
    ('E14:F16', 14, 5),   # 7
    ('B14:C16', 14, 2),   # 8
    ('K11:L13', 11, 11),  # 9
    ('H11:I13', 11, 8),   # 10
    ('E11:F13', 11, 5),   # 11
    ('B11:C13', 11, 2),   # 12
    ('K8:L10', 8, 11),    # 13
    ('H8:I10', 8, 8),     # 14
    ('E8:F10', 8, 5),     # 15
    ('B8:C10', 8, 2),     # 16
    ('K5:L7', 5, 11),     # 17
    ('H5:I7', 5, 8),      # 18
    ('E5:F7', 5, 5),      # 19
    ('B5:C7', 5, 2)       # 20
]

# 자리 배치
for merge_range, row, col in seat_positions:
    if current_seat <= n:
        x1.merge_cells(merge_range)
        x1.cell(row = row, column = col).value = current_seat
        for row in x1[merge_range]:
            for cell in row:
                cell.border = Thin_border
        current_seat += 1

# 추가 데이터 입력
x1['B22'] = "1-3"
x1['B22'].font = Pretendard

x1['B23'] = "강진아"
x1['B23'].font = Pretendard


# 엑셀 파일 저장
xlsx.save("test.xlsx")